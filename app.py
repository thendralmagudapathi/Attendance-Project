import pandas as pd
from unicodedata import name
from attr import validate
from django.shortcuts import redirect
from flask import Flask,render_template,request,flash, url_for
import os
from os import listdir
from PIL import Image
from numpy import asarray
from numpy import expand_dims
from matplotlib import pyplot
from keras.models import load_model
import numpy as np
from mtcnn.mtcnn import MTCNN
import pickle
import cv2
from openpyxl import Workbook
import datetime
import keras_facenet
#pip isntallfrom sklearn.datasets import fetch_rcv1

book = Workbook()
sheet = book.active


app=Flask(__name__,template_folder="C:/Users/thend/Simple-Smart-Attendance-main/main/templates")
app.secret_key="123"

app.config['UPLOAD_FOLDER']="C:/Users/thend/Simple-Smart-Attendance-main/main/static/images/"

detector = MTCNN()
model_path ="C:/Users/thend/Simple-Smart-Attendance-main/main/facenet_keras.h5"
MyFaceNet = load_model(model_path,compile=False)
facenet = keras_facenet.FaceNet('20180402-114759')

#importing trained models
myfile = open("C:/Users/thend/Simple-Smart-Attendance-main/main/data.pkl", "rb")
database = pickle.load(myfile)
myfile.close()


#saving the results as present and absent in workbook
def excel_sheet(output):
    total_stud_folder='C:/Users/thend/Simple-Smart-Attendance-main/Train_Images/'
    students_name_exl = set()
    for folder_name in listdir(total_stud_folder):
        students_name_exl.add(folder_name)




    attendance_exl = []
    for names in students_name_exl:
        attendance_exl.append(names)
    attendance_exl

    now= datetime.datetime.now()
    today=now.day
    month=now.month
    year = now.year

    #index in excel
    sheet["A1"] = "Students_name"
    for row in range(1,30):
        if row <= len(attendance_exl):
            sheet.cell(row=row+1,column=1).value=attendance_exl[row-1]

    #casting outputs to Type List
    list_ouput = []
    for i in output:
        list_ouput.append(i)
    list_ouput

    #markings in excel
    markings = {}
    for i in range(len(attendance_exl)):
        if attendance_exl[i] not in list_ouput:
            markings[i]="absent"
        else:
            markings[i]="present"
    #marking values in second column with date

    sheet["B1"] = "{}/{}/{}".format(today,month,year)
    markings_list = []
    for i in markings:
        markings_list.append(markings[i])
        
    sheet["A1"] = "Students_name"
    for row in range(1,30):
        if row <= len(markings_list):
            sheet.cell(row=row+1,column=2).value=markings_list[row-1]
    book.save("C:/Users/thend/Simple-Smart-Attendance-main/main/Excel_Result/stud.xlsx") 



def findFaces(data):
    gbr1= cv2.imread(data)                                  #(use this for video capturing)js_to_image(data)
    gbr = cv2.cvtColor(gbr1, cv2.COLOR_BGR2RGB)
    gbr = Image.fromarray(gbr)                  # konversi dari OpenCV ke PIL
    gbr_array = asarray(gbr)

    wajah = detector.detect_faces(gbr1)
    final_set = set()
#     extracted_faces = 0
    
    for result in wajah:
        x1, y1, w, h = result['box']
    #for (x1,y1,w,h) in wajah:
        x1, y1 = abs(x1), abs(y1)
        x2, y2 = x1 + w, y1 + h

        face = gbr_array[y1:y2, x1:x2]
#         #for extracting images from output
        
#         cv2.imshow("detected.jpg",face)
        
#         cv2.imwrite("c:/Users/nelso/FaceRecognition_main/My_project/main/Extracted/"+str(extracted_faces)+".jpg", gbr1)
#         extracted_faces+=1


        face = Image.fromarray(face)                       
        face = face.resize((160,160))
        face = asarray(face)

        face = face.astype('float32')
        mean, std = face.mean(), face.std()
        face = (face - mean) / std

        face = expand_dims(face, axis=0)
        signature = MyFaceNet.predict(face)

        min_dist=100
        identity=' '
        for key, value in database.items() :
            
            dist = np.linalg.norm(value-signature)

            if dist < min_dist:
                min_dist = dist
                identity = key
            text=""
            numbers=""
            res=[]
            for i in identity:
                if(i.isdigit()):
                    numbers+=i
                else:
                    text+=i
            res.append(text)
            res.append(numbers)
        final_set.add(text)
        cv2.putText(gbr1,identity, (x1,y1),cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 0), 1, cv2.LINE_AA)
        cv2.rectangle(gbr1,(x1,y1),(x2,y2), (0,255,0), 2)
        print(identity)
       
#print(final_set)


      #EXTRACTING NAMES FROM OUTPUT
    return excel_sheet(final_set)

# reading the data in the csv file
df = pd.read_excel(r"C:/Users/thend/Simple-Smart-Attendance-main/main/Excel_Result/stud.xlsx")


@app.route("/",methods=['GET'])
def home():
        return render_template("Loginpage.html")
    



@app.route("/predict",methods=['GET','POST'])
def upload_predict():
    if request.method=='POST':
        upload_image=request.files['upload_image']

        if upload_image.filename!='':
            filepath=os.path.join(app.config["UPLOAD_FOLDER"],upload_image.filename)
            upload_image.save(filepath)
            flash("File Upload Successfully","success")
            print(filepath)
            print(type(filepath))
            str_path = str(filepath)
            pred = findFaces(str_path)
            path=filepath
            return render_template('upload.html',data=str_path)

    return render_template("upload.html")



@app.route("/table",methods=['GET','POST'])
def table():
    data = pd.read_excel(r"C:/Users/thend/Simple-Smart-Attendance-main/main/Excel_result/stud.xlsx")
    return render_template('table.html',tables=[data.to_html()],titles=[""])


if __name__ == '__main__':
    app.run(debug=True)